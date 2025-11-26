# daliy_copy_task.py
import openpyxl
from datetime import datetime, date
# 🔑 新增這行 import
from openpyxl.cell.cell import MergedCell

def get_cell_value(ws, cell_address):
    """安全讀取單一儲存格的值"""
    try:
        return ws[cell_address].value
    except:
        return None

def find_date_column(ws, row_idx, target_date):
    """在指定列尋找符合 target_date 的欄位索引"""
    max_col = ws.max_column
    for col in range(1, max_col + 1):
        cell_val = ws.cell(row=row_idx, column=col).value
        
        if isinstance(cell_val, datetime):
            cell_val = cell_val.date()
        elif isinstance(cell_val, str):
            try:
                cell_val = datetime.strptime(cell_val, "%Y/%m/%d").date()
            except:
                pass
        
        if cell_val == target_date:
            return col
    return None

def copy_by_mapping_openpyxl(wb_src, wb_dst, tasks):
    """執行 tasks 列表中的所有複製任務"""
    logs = []
    success_count = 0
    fail_count = 0
    
    for idx, task in enumerate(tasks):
        task_label = f"Task {idx+1}"
        
        try:
            # 1. 解析來源 Sheet
            src_sheet_name = task["src_sheet"]
            ws_src = None
            if src_sheet_name in wb_src.sheetnames:
                ws_src = wb_src[src_sheet_name]
            else:
                for name in wb_src.sheetnames:
                    if name in src_sheet_name or src_sheet_name.replace("模板", "") in name:
                        ws_src = wb_src[name]
                        break
            
            if ws_src is None:
                logs.append(f"⚠️ {task_label}: 找不到來源工作表 '{src_sheet_name}'")
                fail_count += 1
                continue

            # 2. 獲取來源日期
            src_date_val = get_cell_value(ws_src, task["src_date_cell"])
            if isinstance(src_date_val, datetime):
                src_date_val = src_date_val.date()
            
            if not src_date_val:
                logs.append(f"⚠️ {task_label}: 無法從 {task['src_date_cell']} 讀取日期")
                fail_count += 1
                continue

            # 3. 讀取來源資料
            src_range_cells = ws_src[task["src_value_range"]]
            src_values = [row[0].value for row in src_range_cells]

            # 4. 解析目的 Sheet
            dst_sheet_name = task["dst_sheet"]
            if dst_sheet_name not in wb_dst.sheetnames:
                logs.append(f"⚠️ {task_label}: 找不到目的工作表 '{dst_sheet_name}'")
                fail_count += 1
                continue
            
            ws_dst = wb_dst[dst_sheet_name]
            
            # 5. 在目的檔尋找對應的日期欄位
            date_row = task["dst_date_row"]
            target_col_idx = find_date_column(ws_dst, date_row, src_date_val)
            
            if not target_col_idx:
                logs.append(f"⚠️ {task_label}: 在 '{dst_sheet_name}' 第 {date_row} 列找不到日期 {src_date_val}")
                fail_count += 1
                continue

            # 6. 計算寫入位置
            dst_start_col = target_col_idx + task["dst_value_start_offset_col"]
            dst_start_row = date_row + task["dst_value_start_offset_row"]
            
            # 7. 執行寫入 (加入 MergedCell 防呆)
            for i, val in enumerate(src_values):
                current_row = dst_start_row + i
                current_col = dst_start_col
                
                dst_cell = ws_dst.cell(row=current_row, column=current_col)
                
                # 🛑 關鍵修正：檢查是否為合併儲存格
                if isinstance(dst_cell, MergedCell):
                    continue # 跳過唯讀格

                dst_cell.value = val
                
            success_count += 1

        except Exception as e:
            logs.append(f"❌ {task_label} 發生錯誤: {str(e)}")
            fail_count += 1

    summary = f"✅ Step 2 彙總：成功 {success_count} 項，失敗 {fail_count} 項。"
    logs.append(summary)
    
    return True, logs