# -*- coding: utf-8 -*-
from pathlib import Path
from openpyxl import load_workbook
import win32com.client as win32
from win32com.client import constants
from datetime import datetime, date
import re
import sys
import time

BASE_DIR = Path(r"C:\Project\daily")

# ========= 公用函式 =========

def find_file(prefix: str) -> Path | None:
    """在資料夾中尋找開頭符合的檔案"""
    for p in BASE_DIR.iterdir():
        if p.is_file() and p.stem.startswith(prefix) and p.suffix.lower() == ".xlsx":
            return p.resolve()
    return None


def col_letter(n: int) -> str:
    """將欄位數字轉為 Excel 字母"""
    s = ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def to_date(v):
    """將 Excel 取出的日期值統一成 date 物件"""
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return v


def print_progress_bar(current, total, prefix="", length=40):
    """在終端機顯示進度條"""
    percent = current / total
    filled = int(length * percent)
    bar = "█" * filled + "-" * (length - filled)
    sys.stdout.write(f"\r{prefix} |{bar}| {percent*100:5.1f}%")
    sys.stdout.flush()
    if current == total:
        sys.stdout.write("\n")


# ========= 找日期對應欄位 =========

def find_column_letter():
    daily = find_file("114年dailyTool-單日")
    report = find_file("影視業務日報表")
    if not daily or not report:
        print("找不到指定開頭的檔案")
        return None

    wb_daily = load_workbook(daily, data_only=True)
    wb_report = load_workbook(report, data_only=True)

    j2_value = wb_report["摘要表"]["J2"].value
    if j2_value != 1:
        print("⚠️ 天數 > 1 → 改為引用 NH 欄")
        return "NH"  # 直接指定替換為 NH 欄位

    target_date = to_date(wb_daily["日統計模板"]["B1"].value)
    ws_stat = wb_report["日統計"]
    row_values = [cell.value for cell in ws_stat[56]]

    for idx, val in enumerate(row_values, start=1):
        if to_date(val) == target_date:
            col = col_letter(idx)
            print(f"✅ 日期 {target_date} 位於 {col} 欄")
            return col

    print("⚠️ 找不到對應日期")
    return None


# ========= 多工作表公式替換 =========

def replace_formulas_multi_sheet(col_letter: str):
    """在多個工作表、多個區間替換公式中的日統計欄位（含進度條）"""
    report = find_file("影視業務日報表")
    if not report or not col_letter:
        print("❌ 找不到檔案或欄位")
        return

    excel = win32.DispatchEx("Excel.Application")
    try:
       excel.Visible = False
       excel.ScreenUpdating = False
       excel.DisplayAlerts = False
    except Exception:
        pass
    
    try:
        wb = excel.Workbooks.Open(str(report))
        try:
            excel.Calculation = constants.xlCalculationManual
        except Exception:
            pass

        pattern = re.compile(r"(日統計!)\$?[A-Z]+\$?(\d+)")
        targets = {
            "MOD(含無上網)": [
                "D6:D23", "H6:H23", "L6:L23", "P6:P23", "T6:T23",
                "D29:D46", "H29:H46", "L29:L46", "P29:P46", "T29:T46",
                "D52:D69", "H52:H69", "L52:L69", "P52:P69", "T52:T69",
            ],
            "無上網MOD": [
                "D6:D23", "H6:H23", "L6:L23", "P6:P23", "T6:T23",
                "D29:D46", "H29:H46", "L29:L46", "P29:P46", "T29:T46",
                "D52:D69", "H52:H69", "L52:L69", "P52:P69", "T52:T69",
            ],
        }

        # 計算總格數
        total_cells = sum(
            ws.Range(rng).Rows.Count * ws.Range(rng).Columns.Count
            for ws_name in targets
            for rng in targets[ws_name]
            for ws in [wb.Worksheets(ws_name)]
        )

        total_changes = 0
        processed = 0
        start_time = time.time()

        for sheet_name, ranges in targets.items():
            ws = wb.Worksheets(sheet_name)
            for rng in ranges:
                arr = ws.Range(rng).Formula
                new_arr = []
                for row in arr:
                    new_row = []
                    for f in row:
                        if f and isinstance(f, str) and "日統計!" in f:
                            # 直接改成指向 '日統計'!NH
                            new_formula = pattern.sub(rf"\1{col_letter}\2", f)
                            new_row.append(new_formula)
                            total_changes += 1
                        else:
                            new_row.append(f)
                        processed += 1
                    new_arr.append(new_row)
                ws.Range(rng).Formula = new_arr
                print_progress_bar(processed, total_cells, prefix=f"處理中")

        wb.Save()
        elapsed = time.time() - start_time

    finally:
        try:
            excel.Calculation = constants.xlCalculationAutomatic
        except Exception:
            pass
        excel.ScreenUpdating = True
        wb.Close(SaveChanges=True)
        excel.Quit()


# ========= 主流程 =========

if __name__ == "__main__":
    col = find_column_letter()  # 若 J2 != 1 則 col = "NH"
    replace_formulas_multi_sheet(col)
