# -*- coding: utf-8 -*-
from pathlib import Path
import win32com.client as win32
from openpyxl.utils import get_column_letter
import psutil
import time
import sys

BASE_DIR = Path(r"C:\Project\daily")


# =========================================================
# 進度條
# =========================================================
def print_progress_bar(current, total, prefix="", length=40):
    """顯示進度條動畫（█ + 百分比）"""
    percent = current / total if total else 1
    filled = int(length * percent)
    bar = "█" * filled + "-" * (length - filled)
    sys.stdout.write(f"\r{prefix} |{bar}| {percent*100:5.1f}%")
    sys.stdout.flush()
    if current == total:
        sys.stdout.write("\n")


# =========================================================
# 清殘餘 Excel
# =========================================================
def kill_excel_zombies():
    for proc in psutil.process_iter(attrs=["name"]):
        if proc.info["name"] and "EXCEL" in proc.info["name"].upper():
            try:
                proc.kill()
            except Exception:
                pass


# =========================================================
# 找檔案
# =========================================================
def find_file(prefix: str) -> Path | None:
    for p in BASE_DIR.iterdir():
        if p.is_file() and p.stem.startswith(prefix) and p.suffix.lower() in (".xlsx", ".csv"):
            return p.resolve()
    return None


# =========================================================
# 主程式（含進度條）
# =========================================================
def main():

    kill_excel_zombies()
    time.sleep(1)

    # === 啟動 Excel ===
    excel = win32.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.ScreenUpdating = False

    # === 找檔案 ===
    print_progress_bar(1, 7, prefix="總進度")
    mod_file = find_file("mod_unrent_unfinish")
    daily_file = find_file("114年dailyTool-單日")
    report_file = find_file("影視業務日報表")

    if not all([mod_file, daily_file, report_file]):
        print("❌ 有檔案找不到，請確認資料夾內容：", BASE_DIR)
        if not mod_file: print("→ 缺少 mod_unrent_unfinish")
        if not daily_file: print("→ 缺少 114年dailyTool-單日")
        if not report_file: print("→ 缺少 影視業務日報表")
        excel.Quit()
        return


    # === Step 0: 開檔進度 ===
    print_progress_bar(2, 7, prefix="總進度")
    wb_mod = excel.Workbooks.Open(str(mod_file))
    wb_daily = excel.Workbooks.Open(str(daily_file))
    wb_report = excel.Workbooks.Open(str(report_file))
    ws_report1 = wb_report.Worksheets("工作表1")

    # ============================
    # Step 1: A3:E21 → I53:M71
    # ============================
    print_progress_bar(3, 7, prefix="總進度")

    try:
        src = wb_mod.Worksheets(1).Range("A3:E21")
        ws_report1.Range("I53:M71").Value = src.Value
    except Exception as e:
        print("⚠️ Step 1 發生錯誤：", e)


    # ============================
    # Step 2: J7:J30 → B3:B26
    # ============================
    print_progress_bar(4, 7, prefix="總進度")
    try:
        src = wb_daily.Worksheets("DAY1").Range("J7:J30")
        ws_report1.Range("B3:B26").Value = src.Value
    except Exception as e:
        print("⚠️ Step 2 發生錯誤：", e)


    # ============================
    # Step 3: 待拆數 (首列)
    # ============================
    print_progress_bar(5, 7, prefix="總進度")

    # --- 優化：將重複使用的物件變數化 ---
    ws_template = wb_daily.Worksheets("日統計模板")
    ws_pending = wb_report.Worksheets("待拆數")
    key_val = ws_template.Range("B1").Value

    try:
        ws_template = wb_daily.Worksheets("日統計模板")
        key_val = ws_template.Range("B1").Value
        ws_pending = wb_report.Worksheets("待拆數")

        last_col = ws_pending.Cells(1, ws_pending.Columns.Count).End(-4159).Column

        # 比對欄名進度條
        for c in range(1, last_col + 1):
            print_progress_bar(c, last_col, prefix="  比對欄位中")
            if ws_pending.Cells(1, c).Value == key_val:
                col_letter = get_column_letter(c)
                src = ws_report1.Range("M31:M48")
                ws_pending.Range(f"{col_letter}2:{col_letter}19").Value = src.Value
                break
        else:
            print("⚠️ Step 3：找不到對應欄位名稱")
        # --- 優化：一次讀取整列資料，在記憶體中比對 ---
        header_values = ws_pending.Range(ws_pending.Cells(1, 1), ws_pending.Cells(1, last_col)).Value[0]
        
        col_index = None
        try:
            # .index() 在 Python 中查找速度極快
            col_index = header_values.index(key_val) + 1
        except (ValueError, TypeError):
            print(f"⚠️ Step 3：在 '待拆數' 工作表第一列找不到符合 '{key_val}' 的欄位")

        if col_index:
            col_letter = get_column_letter(col_index)
            src_values = ws_report1.Range("M31:M48").Value
            ws_pending.Range(f"{col_letter}2:{col_letter}19").Value = src_values

    except Exception as e:
        print("⚠️ Step 3 發生錯誤：", e)


    # ============================
    # Step 4: 待拆數 第23行比對
    # ============================
    print_progress_bar(6, 7, prefix="總進度")

    try:
        ws_template = wb_daily.Worksheets("日統計模板")
        key_val = ws_template.Range("B1").Value
        ws_pending = wb_report.Worksheets("待拆數")

        last_col = ws_pending.Cells(23, ws_pending.Columns.Count).End(-4159).Column

        # 欄位進度條
        for c in range(1, last_col + 1):
            if ws_pending.Cells(23, c).Value == key_val:
                col_letter = get_column_letter(c)
                src = ws_report1.Range("N31:N48")
                ws_pending.Range(f"{col_letter}24:{col_letter}41").Value = src.Value
                break
        else:
            print("⚠️ Step 4：找不到欄位名稱")
        # --- 優化：同樣一次讀取整列資料 ---
        header_values_row23 = ws_pending.Range(ws_pending.Cells(23, 1), ws_pending.Cells(23, last_col)).Value[0]

        col_index = None
        try:
            col_index = header_values_row23.index(key_val) + 1
        except (ValueError, TypeError):
            print(f"⚠️ Step 4：在 '待拆數' 工作表第 23 列找不到符合 '{key_val}' 的欄位")

        if col_index:
            col_letter = get_column_letter(col_index)
            src_values = ws_report1.Range("N31:N48").Value
            ws_pending.Range(f"{col_letter}24:{col_letter}41").Value = src_values

    except Exception as e:
        print("⚠️ Step 4 發生錯誤：", e)


    # ============================
    # Step X: 完成 & 關閉
    # ============================
    print_progress_bar(7, 7, prefix="總進度")
    wb_report.Save()
    wb_mod.Close(False)
    wb_daily.Close(False)
    wb_report.Close(True)
    excel.Quit()

    print("\n🎉 完成待拆工作表更新！")


# =========================================================
# RUN
# =========================================================
if __name__ == "__main__":
    main()
